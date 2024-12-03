from .utility import log_execution_time, list_blob_files

# from .utility import log_execution_time
import os
import requests
from dotenv import load_dotenv
import logging
import hashlib

from openpyxl import load_workbook  # Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    PatternFill,
    Alignment,
    Font,
    NamedStyle,
    Protection,
    # Font,
    # PatternFill,
    # Color,
)

# from openpyxl import load_workbook  # Workbook,

# from openpyxl.styles import NamedStyle, Font, PatternFill, Color
from PIL import Image

# import cairosvg

from google.cloud import storage

# import xlwings as xw

load_dotenv()
LOGO_DIR = os.getenv("LOGO_DIR")
project = os.getenv("PROJECT_ID")
# bucket_name = os.getenv("BUCKET_ID")
storage_client = storage.Client(project)
bucket = storage_client.bucket(LOGO_DIR)
default_logo_path = "/app/static/contents/avocat/default.png"

# Set a password for all sheets
worksheet_password = "test"

title_cells = [
    ("ACCUEIL", "F4:N4"),
    ("SCORING", "H9:I9"),
    ("SCORING", "H12:I12"),
    ("SIMULATEUR_2", "F3:G3"),
    ("ANALYSE SIMULATIONS", "D2:E2"),
    ("ANALYSE SIMULATIONS", "G2:J2"),
    ("ANALYSE SIMULATIONS", "L2:R2"),
    ("SIMULATEUR", "F8:H8"),
    ("SIMULATEUR", "F14:H14"),
    ("SIMULATEUR", "F20:H20"),
    ("LIENS UTILES", "C4:L4"),
    ("LIENS UTILES", "C14:L14"),
]

attribute_cells_centered = [
    ("ACCUEIL", "F5:N6"),
    ("ACCUEIL", "F7:N16"),
    ("ACCUEIL", "F18:N21"),
    ("SCORING", "H10:I10"),
    ("SCORING", "H13:I13"),
    ("SIMULATEUR_2", "B12:H12"),
]

attribute_cells_left = [
    ("SIMULATEUR_2", "C3:C9"),
    ("SIMULATEUR_2", "F4:F9"),
    ("ANALYSE SIMULATIONS", "D3:D10"),
    ("ANALYSE SIMULATIONS", "G3:I10"),
    ("ANALYSE SIMULATIONS", "L3:P10"),
    ("SIMULATEUR", "F9:G12"),
    ("SIMULATEUR", "F15:G18"),
    ("SIMULATEUR", "F22:G25"),
    ("LIENS UTILES", "C8:F10"),
    ("LIENS UTILES", "C17:F19"),
    ("LIENS UTILES", "C21:F23"),
    ("LIENS UTILES", "C25:F27"),
    ("LIENS UTILES", "I8:L10"),
    ("LIENS UTILES", "I17:L19"),
    ("LIENS UTILES", "I21:L23"),
    ("LIENS UTILES", "I25:L27"),
]


background = [
    ("ACCUEIL", "A1:B400"),
    ("ACCUEIL", "C1:Q2"),
    ("ACCUEIL", "R1:II400"),
    ("ACCUEIL", "C29:Q4000"),
    ("SCORING", "A1:F400"),
    ("SCORING", "G1:J6"),
    ("SCORING", "K1:II400"),
    ("SCORING", "G16:J400"),
    ("VOLUMES_PIECES", "A1:II400"),
    ("VOLUMES_SURFACES", "A1:II400"),
    ("PRIX_PIECES", "A1:II400"),
    ("SIMULATEUR_2", "B1:H1"),
    ("SIMULATEUR_2", "B11:H11"),
    ("SIMULATEUR_2", "A1:A400"),
    ("SIMULATEUR_2", "I1:II400"),
    ("SIMULATEUR_2", "B113:H400"),
    ("ANALYSE SIMULATIONS", "D1:R1"),
    ("ANALYSE SIMULATIONS", "D10:R10"),
    ("ANALYSE SIMULATIONS", "D29:R400"),
    ("ANALYSE SIMULATIONS", "A1:C400"),
    ("ANALYSE SIMULATIONS", "S1:II400"),
    ("ANALYSE SIMULATIONS", "F2:F9"),
    ("ANALYSE SIMULATIONS", "K2:K9"),
    ("SIMULATEUR", "A1:D400"),
    ("SIMULATEUR", "J1:II400"),
    ("SIMULATEUR", "E1:I5"),
    ("SIMULATEUR", "E28:I400"),
    ("LIENS UTILES", "A1:A400"),
    ("LIENS UTILES", "N1:II400"),
    ("LIENS UTILES", "B1:M1"),
    ("LIENS UTILES", "B30:M400"),
]

text_cells = [
    ("ACCUEIL", "J23:J26"),
    ("SCORING", "G17:J17"),
    ("VOLUMES_PIECES", "B27:V27"),
    ("VOLUMES_SURFACES", "B27:V27"),
    ("PRIX_PIECES", "B21:V21"),
]

text_cells = [
    ("ACCUEIL", "J23:J26"),
    ("SCORING", "G17:J17"),
    ("VOLUMES_PIECES", "B27:V27"),
    ("VOLUMES_SURFACES", "B27:V27"),
    ("PRIX_PIECES", "B21:V21"),
]

num_cells_euro = [
    ("SIMULATEUR_2", "D3:D5"),
    ("SIMULATEUR_2", "D8:D9"),
    ("SIMULATEUR_2", "G4:G7"),
    ("SIMULATEUR_2", "B13:F112"),
    ("ANALYSE SIMULATIONS", "E4:E4"),
    ("ANALYSE SIMULATIONS", "E6:E6"),
    ("ANALYSE SIMULATIONS", "E8:E8"),
    ("ANALYSE SIMULATIONS", "J4:J4"),
    ("ANALYSE SIMULATIONS", "J6:J6"),
    ("ANALYSE SIMULATIONS", "J8:J8"),
    ("ANALYSE SIMULATIONS", "Q4:R4"),
    ("ANALYSE SIMULATIONS", "Q6:R6"),
    ("ANALYSE SIMULATIONS", "Q8:R8"),
    ("SIMULATEUR", "H9:H12"),
    ("SIMULATEUR", "H15:H18"),
    ("SIMULATEUR", "H22:H23"),
]

num_cells_percent = [
    ("SIMULATEUR_2", "D6:D7"),
    ("SIMULATEUR_2", "G13:H112"),
    ("ANALYSE SIMULATIONS", "E5:E5"),
    ("ANALYSE SIMULATIONS", "E7:E7"),
    ("ANALYSE SIMULATIONS", "E9:E9"),
    ("ANALYSE SIMULATIONS", "J5:J5"),
    ("ANALYSE SIMULATIONS", "J7:J7"),
    ("ANALYSE SIMULATIONS", "J9:J9"),
    ("ANALYSE SIMULATIONS", "Q5:R5"),
    ("ANALYSE SIMULATIONS", "Q7:R7"),
    ("ANALYSE SIMULATIONS", "Q9:R9"),
    ("SIMULATEUR", "H24:H25"),
]

comments = [
    ("SIMULATEUR_2", "E3:E3"),
    ("SIMULATEUR_2", "E4:E4"),
    ("SIMULATEUR_2", "E5:E5"),
    ("SIMULATEUR_2", "E6:E6"),
    ("SIMULATEUR_2", "E7:E7"),
    ("SIMULATEUR_2", "E8:E8"),
    ("SIMULATEUR_2", "E9:E9"),
]


def hash_email_md5(email):
    # Convert the email to bytes and hash it using MD5
    return hashlib.md5(email.encode()).hexdigest()


@log_execution_time
def upload_logo(logo, logoLink, user):

    if user.logo:
        # Check if a blob containing the user's email exists and delete it
        currentLogo = user.logo.split("/")
        currentLogo = currentLogo[-1]
        logging.debug(f"currentLogo: {currentLogo}.")
        existing_blob = bucket.blob(currentLogo)
        logging.debug(f"currentLogo EXIST: {existing_blob.exists()}")
        if existing_blob.exists():
            existing_blob.delete()  # Remove the existing blob
            logging.debug("Remove logo file in GCS.")

    if logo:
        logging.debug(f"Upload logo file: {logo}.")
        # Create a unique filename using the user's email
        filename = f"{hash_email_md5(user.email)}.{logo.filename.split('.')[-1]}"
        # filename = filename.replace(" ", "_")
        # filename = os.path.join(LOGO_DIR, filename)
        logging.debug(f"Upload logo filename: {filename}.")
        # Replace @ and . for valid filename
        blob = bucket.blob(filename)
        blob.upload_from_file(logo)  # Pass the file object
        # Return the public URL or the blob name
        logging.debug("Upload logo file in GCS.")
        # return blob.public_url  # or blob.name for internal use
        # Construct the authenticated URL manually
        authenticated_url = (
            f"https://storage.cloud.google.com/{bucket.name}/{blob.name}"
        )

        # Return the authenticated URL
        return authenticated_url

    # elif logoLink:
    #     logging.debug("Return logoLink.")
    #     return logoLink


@log_execution_time
def reshape_image(image, max_width=125, max_height=125):
    # Calculate the aspect ratio
    aspect_ratio = image.width / image.height

    # Resize while maintaining the aspect ratio
    if image.width > image.height:
        image.width = max_width
        image.height = max_width / aspect_ratio
    else:
        image.height = max_height
        image.width = max_height * aspect_ratio

    return image


@log_execution_time
def convert_to_png(input_image_path):
    format = input_image_path.split(".")[-1]
    if format == ".png":
        return input_image_path
    output_image_path = input_image_path.split(".")[0] + ".png"
    with Image.open(input_image_path) as img:
        # Convert the image to RGB mode (if not already)
        img = img.convert("RGB")
        # Save the image as PNG
        img.save(output_image_path, "PNG")
    return output_image_path


# def convert_svg_to_png(svg_path, png_path):
#     # Convert SVG to PNG
#     cairosvg.svg2png(url=svg_path, write_to=png_path)
#     # return png_path


# def convert_image(input_image_path):
#     # Convert SVG to PNG
#     format = input_image_path.split(".")[-1]
#     output_path = input_image_path.split(".")[0] + ".png"
#     if format == "svg":
#         convert_svg_to_png(input_image_path, output_path)
#     else:
#         convert_to_png(input_image_path, output_path)
#     return output_path


@log_execution_time
def add_user_logo(wb, current_user, report_dir, bucket):
    # Load the existing workbook
    # wb = load_workbook(report_path)

    # Select the 'ACCUEIL' sheet
    ws = wb["ACCUEIL"]
    # logging.debug(f"REPORT PATH: {report_path}")

    if current_user.logo:
        # Extract the image filename
        image_filename = current_user.logo.split("/")[-1]
        logo_path = "/app/" + report_dir + "/" + image_filename
        logging.debug(f"LOGO PATH: {logo_path}")

        # Check if the logo URL contains "http" indicating it's a web URL
        if "storage" not in current_user.logo:
            # Download the logo directly from the URL
            response = requests.get(current_user.logo)
            if response.status_code == 200:
                with open(logo_path, "wb") as f:
                    f.write(response.content)
                logging.debug(
                    f"Succeed to download image from {current_user.logo}. Status code: {response.status_code}"
                )
            else:
                logging.debug(
                    f"Failed to download image from {current_user.logo}. Status code: {response.status_code}"
                )
                return  # Exit the function if download fails
        else:
            # Download the logo from the bucket
            blob = bucket.blob(image_filename)
            blob.download_to_filename(logo_path)

        logging.debug(f"APP FILES: {'| '.join(list_blob_files('/app'))}")
        logging.debug(f"REPORTS DIR FILES: {'| '.join(list_blob_files(report_dir))}")
        # Load and insert the new image at the desired position
        logo_path = convert_to_png(logo_path)
        image = XLImage(logo_path)
    else:
        logo_path = convert_to_png(default_logo_path)
        image = XLImage(logo_path)

    image = reshape_image(image, max_width=175, max_height=175)
    # Insert the new image (overwrite the old one by placing it in the same position)
    ws.add_image(image, "I7")  # Insert the image at cell I3

    bg_color = "FF215867"

    ws.merge_cells("F6:N6")
    fullName = f"{current_user.firstname} {current_user.name}"
    # contact = f"""{contact}
    # {current_user.email}"""

    # ########## fullName
    ws["F6"].value = fullName
    ws["F6"].alignment = Alignment(horizontal="center")
    # Apply the fill to each cell in the range G15:M15
    for row in ws["F6:N6"]:
        for cell in row:
            cell.fill = PatternFill(
                start_color=bg_color, end_color=bg_color, fill_type="solid"
            )

    # ########## email
    ws.merge_cells("F15:N15")
    ws["F15"].value = current_user.email
    ws["F15"].alignment = Alignment(horizontal="center")
    # Apply the fill to each cell in the range G15:M15
    for row in ws["F15:N15"]:
        for cell in row:
            cell.fill = PatternFill(
                start_color=bg_color, end_color=bg_color, fill_type="solid"
            )
    ws["F15"].alignment = Alignment(horizontal="center")

    # ########## website
    if current_user.website:
        ws.merge_cells("F16:N16")
        ws["F16"].value = "Site Web"

        ws["F16"].hyperlink = current_user.website
        ws["F16"].style = (
            "Hyperlink"  # Optional: Change the style to Hyperlink if needed
        )
        # Apply the fill to each cell in the range G16:M16
        ws["F16"].alignment = Alignment(horizontal="center")
        for row in ws["F16:N16"]:
            for cell in row:
                cell.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )
        ws["F16"].font = Font(color="FFFFFF", bold=True)

    # Set the height of row 5 to 40 points
    ws.row_dimensions[10].height = 30  # 40 points
    return wb


@log_execution_time
def assign_style(wb, wsName, cellName, fontFamily, fontColor, bgColor):
    # def add_user_logo(current_user, report_dir, report_path, bucket):
    # Load the existing workbook
    # wb = load_workbook(report_path)

    # Select the 'ACCUEIL' sheet
    ws = wb[wsName]

    # Open or create a workbook
    # wb = load_workbook('example.xlsx')  # Or use: Workbook() to create a new one
    # ws = wb.active

    # Step 1: Define the style
    custom_style = NamedStyle(name="custom_style")
    custom_style.font = Font(name=fontFamily, size=12, color=fontColor)  # White font
    custom_style.fill = PatternFill(
        start_color=bgColor, end_color=bgColor, fill_type="solid"  # Green background
    )

    # Step 2: Register the style
    wb.add_named_style(custom_style)

    # Step 3: Assign style to named cells
    # Assuming the named cell is registered in Excel as 'MyCell'
    cell = ws[cellName]  # Access named cell
    cell.style = "custom_style"

    # Save the workbook
    # wb.save('example.xlsx')
    return wb


@log_execution_time
def apply_custom_style_to_range(
    wb,
    cell_list,
    font_family,
    font_color,
    font_size,
    bg_color,
    alignment,
    bold,
    style_name,
    number_format=None,
):
    """
    Apply a custom style to a specified range of cells.
    Parameters:
        - wb: Workbook object
        - ws_name: Name of the worksheet
        - cell_range: Range of cells (e.g., 'F6:N6')
        - font_family: Font family name (e.g., 'Arial')
        - font_color: Font color in hex (e.g., 'FFFFFF' for white)
        - bg_color: Background color in hex (e.g., '0000FF' for blue)
    """

    unlocked_cells_list = [
        ("SCORING", ["H10", "I10"]),
        ("SIMULATEUR_2", ["D3", "D4", "D5", "D8", "D9", "G4", "G5", "G6", "G7"]),
    ]

    # Define the custom style
    custom_style = NamedStyle(name=style_name)
    # if number_format:
    #     custom_style.font = Font(name=font_family, size=12, color=f"FF{font_color}", bold=bold, number_format=number_format)  # aRGB format
    # else:
    custom_style.font = Font(
        name=font_family, size=font_size, color=f"FF{font_color}", bold=bold
    )  # aRGB format

    custom_style.fill = PatternFill(
        start_color=f"FF{bg_color}", end_color=f"FF{bg_color}", fill_type="solid"
    )
    custom_style.alignment = alignment

    # Register the style if it doesn't already exist
    if style_name not in wb.named_styles:
        wb.add_named_style(custom_style)

    # Iterate through each cell in the specified range
    for info in cell_list:
        # Select the worksheet
        ws = wb[info[0]]
        cell_range = info[1]
        unlocked_cells = list(filter(lambda s: s[0] == info[0], unlocked_cells_list))
        if unlocked_cells:
            unlocked_cells = unlocked_cells[0][1]
            # print(unlocked_cells)
        for row in ws[cell_range]:
            for cell in row:
                cell.style = style_name  # Apply the custom style
                if unlocked_cells and (cell.coordinate in unlocked_cells):
                    cell.protection = Protection(locked=False)
                if number_format:
                    cell.number_format = number_format  # Apply the number format
    return wb


@log_execution_time
def customise_workbook(current_user, report_path, report_dir, bucket_logo):

    wb = load_workbook(report_path)

    wb = add_user_logo(wb, current_user, report_dir, bucket_logo)

    wb = apply_custom_style_to_range(
        wb=wb,
        cell_list=title_cells,  # Range to style
        font_family=current_user.fontFamily,
        font_color=current_user.title_font_color[1:],  # White font
        font_size=12,
        bg_color=current_user.title_color[1:],
        alignment=Alignment(horizontal="center", vertical="center"),
        bold=1,
        style_name="title_style",
    )

    wb = apply_custom_style_to_range(
        wb=wb,
        cell_list=attribute_cells_centered,  # Range to style
        font_family=current_user.fontFamily,
        font_color=current_user.attribut_font_color[1:],  # White font
        font_size=11,
        bg_color=current_user.attribut_color[1:],
        alignment=Alignment(horizontal="center", vertical="center"),
        bold=1,
        style_name="attribute_style_centered",
    )

    wb = apply_custom_style_to_range(
        wb=wb,
        cell_list=attribute_cells_left,  # Range to style
        font_family=current_user.fontFamily,
        font_color=current_user.attribut_font_color[1:],  # White font
        font_size=11,
        bg_color=current_user.attribut_color[1:],
        alignment=Alignment(horizontal="left", vertical="center"),
        bold=1,
        style_name="attribute_style_left",
    )

    wb = apply_custom_style_to_range(
        wb=wb,
        cell_list=background,  # Range to style
        font_family=current_user.fontFamily,
        font_color=current_user.bg_color[1:],  # White font
        font_size=12,
        bg_color=current_user.bg_color[1:],
        alignment=Alignment(horizontal="center", vertical="center"),
        bold=1,
        style_name="bacground_style",
    )

    wb = apply_custom_style_to_range(
        wb=wb,
        cell_list=text_cells,  # Range to style
        font_family=current_user.fontFamily,
        font_color=current_user.bg_color[1:],  # White font
        font_size=11,
        bg_color="ffffff",
        alignment=Alignment(horizontal="center", vertical="center"),
        bold=1,
        style_name="text_style",
    )

    wb = apply_custom_style_to_range(
        wb=wb,
        cell_list=num_cells_euro,  # Range to style
        font_family=current_user.fontFamily,
        font_color=current_user.bg_color[1:],  # White font
        font_size=11,
        bg_color="ffffff",
        alignment=Alignment(horizontal="center", vertical="center"),
        bold=1,
        style_name="num_euro_style",
        number_format="# ##0 €",
    )

    wb = apply_custom_style_to_range(
        wb=wb,
        cell_list=num_cells_percent,  # Range to style
        font_family=current_user.fontFamily,
        font_color=current_user.bg_color[1:],  # White font
        font_size=11,
        bg_color="ffffff",
        alignment=Alignment(horizontal="center", vertical="center"),
        bold=1,
        style_name="num_percent_style",
        number_format="0%",
    )

    wb = apply_custom_style_to_range(
        wb=wb,
        cell_list=comments,  # Range to style
        font_family=current_user.fontFamily,
        font_color="828282",  # White font
        font_size=11,
        bg_color="ffffff",
        alignment=Alignment(horizontal="left", vertical="center"),
        bold=0,
        style_name="comment_style",
    )

    for ws in wb.worksheets:
        # Enable worksheet protection
        ws.protection.sheet = True

        # Set the password for protection
        ws.protection.set_password(worksheet_password)

    wb.save(report_path)
